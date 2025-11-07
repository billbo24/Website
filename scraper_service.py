import io
from openpyxl import Workbook, load_workbook
import PDF_Reader  # assuming this is your custom module

def process_pdfs(uploaded_files):
    """
    Takes a list of uploaded PDFs (Werkzeug FileStorage objects),
    processes them into Excel using PDF_Reader, and returns a BytesIO
    containing the combined Excel.
    """
    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = "Combined Results"

    first = True
    for f in uploaded_files:
        excel_bytes = PDF_Reader.call_everything(f)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not first and i == 1:
                continue
            master_ws.append(row)
        first = False

    output = io.BytesIO()
    master_wb.save(output)
    output.seek(0)
    return output